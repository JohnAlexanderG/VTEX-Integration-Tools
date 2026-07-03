#!/usr/bin/env python3
"""
VTEX Batch Inventory Uploader
-----------------------------
Genera y sube inventario al endpoint batch de VTEX Logistics.

Entrada ERP requerida:
    CODIGO SKU,CODIGO SUCURSAL,EXISTENCIA

Salida VTEX por parte:
    item_id,account_name,container_id,quantity,unlimited,lead_time,supply_date,seller_id
"""

import argparse
import csv
import io
import json
import os
import random
import shutil
import sys
import time
from collections import Counter
from datetime import datetime, timezone
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple
from urllib.request import urlopen

try:
    import requests
except ImportError:
    requests = None

try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = None


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)

ERP_REQUIRED_COLUMNS = ["CODIGO SKU", "CODIGO SUCURSAL", "EXISTENCIA"]
VTEX_BATCH_COLUMNS = [
    "item_id",
    "account_name",
    "container_id",
    "quantity",
    "unlimited",
    "lead_time",
    "supply_date",
    "seller_id",
]
DEFAULT_SKU_REF_COLUMN = "_SKUReferenceCode"
DEFAULT_SKU_ID_COLUMN = "_SkuId"
SKU_REF_COLUMN_ALIASES = [DEFAULT_SKU_REF_COLUMN, "SKU reference code"]
SKU_ID_COLUMN_ALIASES = [DEFAULT_SKU_ID_COLUMN, "SKU ID"]

SUCCESS_FIELDNAMES = [
    "PartNumber",
    "BatchId",
    "PartFile",
    "Rows",
    "Bytes",
    "CreateStatus",
    "UploadStatus",
    "CommitStatus",
    "FinalStatus",
    "DurationSeconds",
]

FAILED_FIELDNAMES = [
    "PartNumber",
    "BatchId",
    "PartFile",
    "Rows",
    "Bytes",
    "Phase",
    "StatusCode",
    "FinalStatus",
    "Error",
    "ErrorFile",
]

SKIPPED_FIELDNAMES = [
    "SourceRow",
    "CodigoSku",
    "CodigoSucursal",
    "Existencia",
    "Reason",
    "Detail",
]

STATE_FIELDNAMES = [
    "PartNumber",
    "BatchId",
    "PartFile",
    "Rows",
    "Bytes",
    "Status",
    "Phase",
    "StatusCode",
    "FinalStatus",
    "Error",
    "Timestamp",
]

SKU_MAP_CONFLICT_FIELDNAMES = [
    "SkuReferenceCode",
    "FirstRow",
    "FirstSkuId",
    "ConflictRow",
    "ConflictSkuId",
]

TERMINAL_SUCCESS = {"DONE", "FINISHED", "COMPLETED", "SUCCESS", "SUCCEEDED", "PROCESSED"}
TERMINAL_FAILURE = {"FAILED", "FAILURE", "ERROR", "CANCELED", "CANCELLED", "ABORTED", "EXPIRED"}
TRANSIENT_STATUS_CODES = {408, 409, 425, 429, 500, 502, 503, 504}
PROGRESS_PREFIX = "__VTEX_JOB_PROGRESS__"


class BatchInventoryError(Exception):
    """Error controlado del uploader batch."""


def emit_progress(payload: Dict[str, Any]) -> None:
    """Emit compact JSONL progress for the webapp while keeping CLI output usable."""
    safe_payload = {"tool_id": "step_67", **payload}
    print(
        f"{PROGRESS_PREFIX}{json.dumps(safe_payload, ensure_ascii=False, separators=(',', ':'))}",
        flush=True,
    )


def load_project_env(env_path: str) -> None:
    """Carga .env con python-dotenv si existe, o con un parser minimo."""
    if load_dotenv:
        load_dotenv(dotenv_path=env_path)
        return

    if not os.path.isfile(env_path):
        return

    with open(env_path, encoding="utf-8", errors="replace") as f:
        for raw_line in f:
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value


def is_nan_like(value: Any) -> bool:
    if value is None:
        return True
    return str(value).strip().lower() in {"nan", "none", "null"}


def strip_float_suffix(value: str) -> str:
    if value.endswith(".0"):
        try:
            number = float(value)
            if number.is_integer():
                return str(int(number))
        except (ValueError, OverflowError):
            return value
    return value


def normalize_sku(value: Any) -> str:
    if is_nan_like(value):
        return ""
    return strip_float_suffix(str(value).strip())


def normalize_warehouse(value: Any, mode: str) -> str:
    if is_nan_like(value):
        return ""
    text = strip_float_suffix(str(value).strip())
    if mode == "zfill3" and text.isdigit() and len(text) < 3:
        return text.zfill(3)
    return text


def parse_non_negative_int(value: Any) -> Optional[int]:
    if is_nan_like(value):
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        number = float(text)
    except (ValueError, OverflowError):
        return None
    if not number.is_integer():
        return None
    integer = int(number)
    if integer < 0:
        return None
    return integer


def parse_sku_id(value: Any) -> Optional[int]:
    if is_nan_like(value):
        return None
    text = strip_float_suffix(str(value).strip())
    if not text:
        return None
    try:
        number = float(text)
    except (ValueError, OverflowError):
        return None
    if not number.is_integer():
        return None
    integer = int(number)
    if integer <= 0:
        return None
    return integer


def bool_text(value: bool) -> str:
    return "true" if value else "false"


def parse_iso_datetime(value: str) -> Optional[datetime]:
    text = str(value or "").strip()
    if not text:
        return None
    if text.endswith("Z"):
        text = f"{text[:-1]}+00:00"
    try:
        parsed = datetime.fromisoformat(text)
    except ValueError:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed


def format_json(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return json.dumps(value, ensure_ascii=False, sort_keys=True)


def validate_required_columns(fieldnames: Sequence[str], required: Sequence[str], label: str) -> None:
    missing = [column for column in required if column not in fieldnames]
    if missing:
        raise BatchInventoryError(
            f"Faltan columnas requeridas en {label}: {', '.join(missing)}\n"
            f"Columnas disponibles: {', '.join(fieldnames)}"
        )


def normalize_column_name(value: str) -> str:
    return str(value or "").strip().casefold()


def resolve_sku_map_column(
    fieldnames: Sequence[str],
    requested_column: str,
    default_column: str,
    aliases: Sequence[str],
    description: str,
    path: str,
) -> str:
    if requested_column in fieldnames:
        return requested_column

    requested_normalized = normalize_column_name(requested_column)
    for fieldname in fieldnames:
        if normalize_column_name(fieldname) == requested_normalized:
            return fieldname

    if requested_column == default_column:
        alias_names = {normalize_column_name(alias) for alias in aliases}
        for fieldname in fieldnames:
            if normalize_column_name(fieldname) in alias_names:
                return fieldname

    available = ", ".join(fieldnames) if fieldnames else "(sin columnas)"
    raise BatchInventoryError(
        f"No se encontro la columna {description} '{requested_column}' en --sku-map: {path}\n"
        f"Columnas disponibles: {available}"
    )


def read_done_part_numbers(state_file: str) -> Set[int]:
    if not os.path.isfile(state_file):
        return set()

    done: Set[int] = set()
    with open(state_file, encoding="utf-8-sig", errors="replace", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                part_number = int(str(row.get("PartNumber", "")).strip())
            except ValueError:
                continue
            if str(row.get("Status", "")).strip().upper() == "DONE":
                done.add(part_number)
    return done


def append_state_row(state_file: str, row: Dict[str, Any], encoding: str) -> None:
    os.makedirs(os.path.dirname(os.path.abspath(state_file)), exist_ok=True)
    should_write_header = not os.path.isfile(state_file) or os.path.getsize(state_file) == 0
    output = {field: row.get(field, "") for field in STATE_FIELDNAMES}
    output["Timestamp"] = output.get("Timestamp") or datetime.now().isoformat(timespec="seconds")

    with open(state_file, "a", encoding=encoding, newline="") as f:
        writer = csv.DictWriter(f, fieldnames=STATE_FIELDNAMES, extrasaction="ignore")
        if should_write_header:
            writer.writeheader()
        writer.writerow(output)


def write_csv(path: str, fieldnames: Sequence[str], rows: Sequence[Dict[str, Any]], encoding: str) -> None:
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    with open(path, "w", encoding=encoding, newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def render_vtex_row(row: Dict[str, Any]) -> str:
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=VTEX_BATCH_COLUMNS, lineterminator="\n")
    writer.writerow(row)
    return buffer.getvalue()


def add_sku_map_rows(
    mappings: Dict[str, str],
    source_rows: Dict[str, int],
    conflicts: List[Dict[str, Any]],
    rows: Iterable[Tuple[int, Dict[str, Any]]],
    ref_column: str,
    id_column: str,
) -> None:
    for index, row in rows:
        sku_ref = normalize_sku(row.get(ref_column, ""))
        if not sku_ref:
            continue
        sku_id = normalize_sku(row.get(id_column, ""))
        if not sku_id:
            sku_id = str(row.get(id_column, "") or "").strip()
        existing = mappings.get(sku_ref)
        if existing is not None and existing != sku_id:
            conflicts.append({
                "SkuReferenceCode": sku_ref,
                "FirstRow": source_rows.get(sku_ref),
                "FirstSkuId": existing,
                "ConflictRow": index,
                "ConflictSkuId": sku_id,
            })
            continue
        mappings[sku_ref] = sku_id
        source_rows.setdefault(sku_ref, index)


def load_sku_map_csv(
    path: str,
    ref_column: str,
    id_column: str,
    encoding: str,
    mappings: Dict[str, str],
    source_rows: Dict[str, int],
    conflicts: List[Dict[str, Any]],
) -> None:
    with open(path, encoding=encoding, errors="replace", newline="") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or []
        resolved_ref_column = resolve_sku_map_column(
            fieldnames,
            ref_column,
            DEFAULT_SKU_REF_COLUMN,
            SKU_REF_COLUMN_ALIASES,
            "referencia SKU",
            path,
        )
        resolved_id_column = resolve_sku_map_column(
            fieldnames,
            id_column,
            DEFAULT_SKU_ID_COLUMN,
            SKU_ID_COLUMN_ALIASES,
            "id SKU",
            path,
        )
        add_sku_map_rows(
            mappings,
            source_rows,
            conflicts,
            enumerate(reader, start=2),
            resolved_ref_column,
            resolved_id_column,
        )


def load_sku_map_xlsx(
    path: str,
    ref_column: str,
    id_column: str,
    header_row: int,
    mappings: Dict[str, str],
    source_rows: Dict[str, int],
    conflicts: List[Dict[str, Any]],
) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise BatchInventoryError("Para leer --sku-map .xlsx instale openpyxl: pip install openpyxl")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        header_values = next(
            sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True),
            None,
        )
        if not header_values:
            raise BatchInventoryError(f"El archivo --sku-map .xlsx no tiene encabezados en la fila {header_row}: {path}")

        fieldnames = []
        for index, value in enumerate(header_values, start=1):
            header = str(value or "").strip()
            fieldnames.append(header if header else f"column_{index}")

        resolved_ref_column = resolve_sku_map_column(
            fieldnames,
            ref_column,
            DEFAULT_SKU_REF_COLUMN,
            SKU_REF_COLUMN_ALIASES,
            "referencia SKU",
            path,
        )
        resolved_id_column = resolve_sku_map_column(
            fieldnames,
            id_column,
            DEFAULT_SKU_ID_COLUMN,
            SKU_ID_COLUMN_ALIASES,
            "id SKU",
            path,
        )

        def iter_data_rows() -> Iterable[Tuple[int, Dict[str, Any]]]:
            for row_number, values in enumerate(
                sheet.iter_rows(min_row=header_row + 1, values_only=True),
                start=header_row + 1,
            ):
                row = {
                    fieldname: values[index] if index < len(values) else ""
                    for index, fieldname in enumerate(fieldnames)
                }
                yield row_number, row

        add_sku_map_rows(
            mappings,
            source_rows,
            conflicts,
            iter_data_rows(),
            resolved_ref_column,
            resolved_id_column,
        )
    finally:
        workbook.close()


def format_sku_map_conflict(conflict: Dict[str, Any]) -> str:
    return (
        f"{conflict.get('SkuReferenceCode')}: filas {conflict.get('FirstRow')} "
        f"y {conflict.get('ConflictRow')} mapean a {conflict.get('FirstSkuId')} "
        f"y {conflict.get('ConflictSkuId')}"
    )


def write_sku_map_conflict_reports(
    output_dir: str,
    timestamp: str,
    sku_map_path: str,
    conflicts: Sequence[Dict[str, Any]],
    encoding: str,
) -> Tuple[str, str]:
    conflicts_csv = os.path.join(output_dir, f"{timestamp}_sku_map_conflicts.csv")
    conflicts_md = os.path.join(output_dir, f"{timestamp}_sku_map_conflicts_REPORT.md")
    write_csv(conflicts_csv, SKU_MAP_CONFLICT_FIELDNAMES, conflicts, encoding)

    unique_refs = sorted({str(conflict.get("SkuReferenceCode", "")) for conflict in conflicts})
    with open(conflicts_md, "w", encoding="utf-8") as f:
        f.write("# Reporte de conflictos en SKU map\n\n")
        f.write(f"**SKU map:** {sku_map_path}\n\n")
        f.write(f"**Conflictos detectados:** {len(conflicts):,}\n\n")
        f.write(f"**Referencias afectadas:** {len(unique_refs):,}\n\n")
        f.write(f"**CSV detalle:** {conflicts_csv}\n\n")
        f.write("## Primeros conflictos\n\n")
        for conflict in conflicts[:50]:
            f.write(f"- {format_sku_map_conflict(conflict)}\n")
        if len(conflicts) > 50:
            f.write(f"- ... {len(conflicts) - 50:,} conflictos adicionales en el CSV\n")

    return conflicts_csv, conflicts_md


def load_sku_map(
    path: str,
    ref_column: str,
    id_column: str,
    encoding: str,
    header_row: int = 2,
    conflict_report_dir: str = "",
    report_timestamp: str = "",
) -> Dict[str, str]:
    if not os.path.isfile(path):
        raise BatchInventoryError(f"No se encontro el archivo --sku-map: {path}")

    mappings: Dict[str, str] = {}
    source_rows: Dict[str, int] = {}
    conflicts: List[Dict[str, Any]] = []

    extension = os.path.splitext(path)[1].lower()
    if extension == ".csv":
        load_sku_map_csv(path, ref_column, id_column, encoding, mappings, source_rows, conflicts)
    elif extension == ".xlsx":
        load_sku_map_xlsx(path, ref_column, id_column, header_row, mappings, source_rows, conflicts)
    else:
        raise BatchInventoryError(f"Formato no soportado para --sku-map: {path}. Use .csv o .xlsx")

    if conflicts:
        conflicted_refs = {str(conflict.get("SkuReferenceCode", "")) for conflict in conflicts}
        for sku_ref in conflicted_refs:
            mappings.pop(sku_ref, None)
            source_rows.pop(sku_ref, None)

        if conflict_report_dir:
            timestamp = report_timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
            conflicts_csv, conflicts_md = write_sku_map_conflict_reports(
                conflict_report_dir,
                timestamp,
                path,
                conflicts,
                encoding,
            )
            print(
                "WARN: Conflictos en --sku-map: "
                f"{len(conflicts):,} conflictos en {len(conflicted_refs):,} referencias. "
                "Se omitiran esas referencias y el flujo continuara.\n"
                f"WARN: Reporte CSV: {conflicts_csv}\n"
                f"WARN: Reporte MD: {conflicts_md}",
                file=sys.stderr,
            )
        else:
            print(
                "WARN: Conflictos en --sku-map: "
                f"{len(conflicts):,} conflictos en {len(conflicted_refs):,} referencias. "
                "Se omitiran esas referencias y el flujo continuara.",
                file=sys.stderr,
            )

    if not mappings:
        raise BatchInventoryError(f"No se cargaron mapeos validos desde --sku-map: {path}")

    return mappings


class VtexBatchClient:
    def __init__(self, account: str, environment: str, app_key: str, app_token: str, timeout: int):
        if requests is None:
            raise BatchInventoryError("Falta instalar requests. Ejecute: pip install requests python-dotenv")

        self.account = account
        self.environment = environment
        self.timeout = timeout
        self.session = requests.Session()
        self.session.headers.update({
            "Accept": "application/json",
            "Content-Type": "application/json",
            "X-VTEX-API-AppKey": app_key,
            "X-VTEX-API-AppToken": app_token,
        })
        env_host = environment
        if env_host.endswith(".com.br"):
            env_host = env_host[:-7]
        self.base_url = f"https://logistics.{env_host}.com.br/{account}/availability/v1/inventory/batch"

    def request_with_retries(self, method: str, url: str, **kwargs: Any) -> Any:
        max_attempts = 5
        for attempt in range(max_attempts):
            try:
                response = self.session.request(method, url, timeout=self.timeout, **kwargs)
            except requests.RequestException as exc:
                if attempt == max_attempts - 1:
                    raise BatchInventoryError(f"request_exception: {exc}")
                self.sleep_before_retry(attempt)
                continue

            if response.status_code not in TRANSIENT_STATUS_CODES or attempt == max_attempts - 1:
                return response
            self.sleep_before_retry(attempt)

        raise BatchInventoryError("No se pudo completar el request")

    @staticmethod
    def sleep_before_retry(attempt: int) -> None:
        base = min(45.0, 0.75 * (2 ** attempt))
        time.sleep(base * (0.85 + random.random() * 0.3))

    def create_batch(self) -> Tuple[int, Dict[str, Any]]:
        response = self.request_with_retries("POST", self.base_url)
        data = response.json() if response.text else {}
        if response.status_code not in (200, 201, 202):
            raise BatchInventoryError(f"create batch fallo HTTP {response.status_code}: {response.text[:1000]}")
        if not isinstance(data, dict):
            raise BatchInventoryError("create batch no devolvio JSON objeto")
        return response.status_code, data

    def commit_batch(self, batch_id: str) -> Tuple[int, Dict[str, Any], str]:
        response = self.request_with_retries("POST", f"{self.base_url}/{batch_id}/commit")
        text = response.text or ""
        data: Dict[str, Any] = {}
        if text:
            try:
                parsed = response.json()
                if isinstance(parsed, dict):
                    data = parsed
            except ValueError:
                data = {}
        return response.status_code, data, text

    def get_status(self, batch_id: str) -> Tuple[int, Dict[str, Any], str]:
        response = self.request_with_retries("GET", f"{self.base_url}/{batch_id}/status")
        text = response.text or ""
        data: Dict[str, Any] = {}
        if text:
            try:
                parsed = response.json()
                if isinstance(parsed, dict):
                    data = parsed
            except ValueError:
                data = {}
        return response.status_code, data, text


def validate_credentials(dry_run: bool) -> Tuple[str, str, str, str]:
    account = os.getenv("VTEX_ACCOUNT_NAME", "")
    environment = os.getenv("VTEX_ENVIRONMENT", "vtexcommercestable")
    app_key = os.getenv("X-VTEX-API-AppKey", "")
    app_token = os.getenv("X-VTEX-API-AppToken", "")

    if dry_run:
        return account or "VTEX_ACCOUNT_NAME", environment, app_key, app_token

    missing = []
    if not account:
        missing.append("VTEX_ACCOUNT_NAME")
    if not app_key:
        missing.append("X-VTEX-API-AppKey")
    if not app_token:
        missing.append("X-VTEX-API-AppToken")
    if missing:
        raise BatchInventoryError(f"Faltan credenciales VTEX en .env: {', '.join(missing)}")
    return account, environment, app_key, app_token


def extract_batch_details(response_data: Dict[str, Any]) -> Tuple[str, str, str, str, Optional[datetime]]:
    batch_id = response_data.get("batchId") or response_data.get("id") or response_data.get("batch_id")
    upload = response_data.get("upload") or {}
    if not isinstance(upload, dict):
        upload = {}

    method = str(upload.get("method") or "PUT").upper()
    url = upload.get("url") or upload.get("uploadUrl") or upload.get("href")
    headers = upload.get("headers") or {}
    if not isinstance(headers, dict):
        headers = {}
    content_type = (
        headers.get("contentType")
        or headers.get("Content-Type")
        or headers.get("content-type")
        or upload.get("contentType")
        or "text/csv"
    )
    expires_at = parse_iso_datetime(str(upload.get("expiresAt") or upload.get("expiration") or ""))

    if not batch_id:
        raise BatchInventoryError(f"create batch no devolvio batchId: {format_json(response_data)[:1000]}")
    if not url:
        raise BatchInventoryError(f"create batch no devolvio upload.url para batch {batch_id}")
    if method != "PUT":
        raise BatchInventoryError(f"Metodo de upload no soportado para batch {batch_id}: {method}")

    return str(batch_id), method, str(url), str(content_type), expires_at


def upload_file_to_presigned_url(path: str, upload_url: str, content_type: str, timeout: int) -> Tuple[int, str]:
    if requests is None:
        raise BatchInventoryError("Falta instalar requests. Ejecute: pip install requests python-dotenv")

    max_attempts = 3
    for attempt in range(max_attempts):
        try:
            with open(path, "rb") as f:
                response = requests.put(
                    upload_url,
                    data=f,
                    headers={"Content-Type": content_type or "text/csv"},
                    timeout=timeout,
                )
        except requests.RequestException as exc:
            if attempt == max_attempts - 1:
                raise BatchInventoryError(f"upload_exception: {exc}")
            VtexBatchClient.sleep_before_retry(attempt)
            continue

        if response.status_code in (200, 204):
            return response.status_code, response.text or ""
        if response.status_code not in TRANSIENT_STATUS_CODES or attempt == max_attempts - 1:
            return response.status_code, response.text or ""
        VtexBatchClient.sleep_before_retry(attempt)

    return 0, "upload_failed"


def status_name(status_data: Dict[str, Any]) -> str:
    for key in ("status", "state", "batchStatus", "processingStatus"):
        value = status_data.get(key)
        if value:
            return str(value).strip().upper()
    return ""


def is_success_status(name: str) -> bool:
    return name.upper() in TERMINAL_SUCCESS


def is_failure_status(name: str) -> bool:
    return name.upper() in TERMINAL_FAILURE


def find_error_url(value: Any, parent_key: str = "") -> Optional[str]:
    if isinstance(value, dict):
        for key, item in value.items():
            key_text = str(key).lower()
            found = find_error_url(item, key_text)
            if found:
                return found
    elif isinstance(value, list):
        for item in value:
            found = find_error_url(item, parent_key)
            if found:
                return found
    elif isinstance(value, str):
        text = value.strip()
        if text.startswith("http") and ("error" in parent_key or "fail" in parent_key):
            return text
    return None


def download_error_file(error_url: str, destination: str, timeout: int) -> str:
    os.makedirs(os.path.dirname(os.path.abspath(destination)), exist_ok=True)
    if requests is not None:
        response = requests.get(error_url, timeout=timeout)
        response.raise_for_status()
        with open(destination, "wb") as f:
            f.write(response.content)
        return destination

    with urlopen(error_url, timeout=timeout) as response:
        with open(destination, "wb") as f:
            shutil.copyfileobj(response, f)
    return destination


def poll_status(
    client: VtexBatchClient,
    batch_id: str,
    poll_interval: float,
    max_wait_minutes: float,
    on_status: Optional[Any] = None,
) -> Tuple[int, Dict[str, Any], str, float]:
    started = time.monotonic()
    deadline = started + max_wait_minutes * 60
    last_code = 0
    last_data: Dict[str, Any] = {}
    last_text = ""
    attempt = 0

    while True:
        attempt += 1
        last_code, last_data, last_text = client.get_status(batch_id)
        name = status_name(last_data)
        elapsed = time.monotonic() - started
        if on_status:
            on_status(last_code, last_data, name, elapsed, attempt)
        if name and (is_success_status(name) or is_failure_status(name)):
            return last_code, last_data, name, elapsed
        if time.monotonic() >= deadline:
            if not name:
                name = "TIMEOUT"
            return last_code, last_data, name, elapsed
        time.sleep(poll_interval)


class PartWriter:
    def __init__(self, output_dir: str, max_bytes: int, encoding: str):
        self.output_dir = output_dir
        self.max_bytes = max_bytes
        self.encoding = encoding
        self.part_number = 0
        self.current_path: Optional[str] = None
        self.current_file: Optional[Any] = None
        self.writer: Optional[csv.DictWriter] = None
        self.rows = 0
        self.bytes_written = 0
        header_text = ",".join(VTEX_BATCH_COLUMNS) + "\n"
        self.header_bytes = len(header_text.encode(self.encoding))

    def open_new_part(self) -> None:
        self.part_number += 1
        self.rows = 0
        self.current_path = os.path.join(self.output_dir, f"pending_part_{self.part_number:04d}.csv")
        self.current_file = open(self.current_path, "w", encoding=self.encoding, newline="")
        self.writer = csv.DictWriter(
            self.current_file,
            fieldnames=VTEX_BATCH_COLUMNS,
            lineterminator="\n",
        )
        self.writer.writeheader()
        header_text = ",".join(VTEX_BATCH_COLUMNS) + "\n"
        self.header_bytes = len(header_text.encode(self.encoding))
        self.bytes_written = self.header_bytes

    def close_current(self) -> Optional[Dict[str, Any]]:
        if not self.current_file or not self.current_path:
            return None
        self.current_file.flush()
        self.current_file.close()
        info = {
            "part_number": self.part_number,
            "path": self.current_path,
            "rows": self.rows,
            "bytes": os.path.getsize(self.current_path),
        }
        self.current_file = None
        self.current_path = None
        self.writer = None
        self.rows = 0
        self.bytes_written = 0
        return info

    def write_row(self, row: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        row_text = render_vtex_row(row)
        row_bytes = len(row_text.encode(self.encoding))
        if self.header_bytes + row_bytes > self.max_bytes:
            return {"row_too_large": True, "row_bytes": row_bytes}

        finalized = None
        if self.current_file is None:
            self.open_new_part()
        elif self.rows > 0 and self.bytes_written + row_bytes > self.max_bytes:
            finalized = self.close_current()
            self.open_new_part()

        if self.writer is None:
            raise BatchInventoryError("No se pudo inicializar writer de parte")
        self.writer.writerow(row)
        self.rows += 1
        self.bytes_written += row_bytes
        return finalized

    def finish(self) -> Optional[Dict[str, Any]]:
        if self.current_file is None:
            return None
        return self.close_current()


class BatchInventoryUploader:
    PHASE_LABELS = {
        "dry_run": "Dry-run",
        "create": "Creando batch",
        "upload": "Subiendo CSV",
        "commit": "Confirmando batch",
        "status_polling": "Consultando status",
        "done": "Completado",
        "failed": "Fallido",
    }
    PHASE_PROGRESS = {
        "dry_run": 100,
        "create": 20,
        "upload": 45,
        "commit": 70,
        "status_polling": 85,
        "done": 100,
        "failed": 100,
    }
    SAFE_STATUS_KEYS = (
        "percent",
        "percentage",
        "progress",
        "processed",
        "processedCount",
        "processedItems",
        "processedRows",
        "total",
        "totalCount",
        "totalItems",
        "totalRows",
        "success",
        "successCount",
        "failed",
        "failedCount",
        "errorCount",
        "errors",
    )

    def __init__(self, args: argparse.Namespace):
        self.args = args
        self.output_dir = os.path.abspath(args.output_dir)
        self.parts_dir = os.path.join(self.output_dir, "parts")
        self.errors_dir = os.path.join(self.output_dir, "errors")
        self.state_file = os.path.join(self.output_dir, "batch_inventory_upload_state.csv")
        self.started_at = datetime.now()
        self.timestamp = self.started_at.strftime("%Y%m%d_%H%M%S")
        self.successful: List[Dict[str, Any]] = []
        self.failed: List[Dict[str, Any]] = []
        self.skipped: List[Dict[str, Any]] = []
        self.reason_counts: Counter = Counter()
        self.parts: List[Dict[str, Any]] = []
        self.phase_durations: Counter = Counter()
        self.rows_read = 0
        self.rows_mapped = 0
        self.parts_started = 0
        self.parts_completed = 0
        self.parts_failed = 0
        self.current_phase = ""
        self.current_batch_id = ""
        self.resume_done_parts = read_done_part_numbers(self.state_file) if args.resume else set()

        os.makedirs(self.output_dir, exist_ok=True)
        os.makedirs(self.parts_dir, exist_ok=True)
        os.makedirs(self.errors_dir, exist_ok=True)

        self.account, self.environment, self.app_key, self.app_token = validate_credentials(args.dry_run)
        self.client: Optional[VtexBatchClient] = None
        if not args.dry_run:
            self.client = VtexBatchClient(
                self.account,
                self.environment,
                self.app_key,
                self.app_token,
                timeout=args.timeout,
            )

    def progress_percent(self, phase: str, final_status: str = "") -> int:
        if phase == "done" or is_success_status(final_status):
            return 100
        if phase == "failed" or is_failure_status(final_status):
            return 100
        return self.PHASE_PROGRESS.get(phase, 0)

    def safe_status_metrics(self, status_data: Dict[str, Any]) -> Dict[str, Any]:
        metrics: Dict[str, Any] = {}
        for key in self.SAFE_STATUS_KEYS:
            value = status_data.get(key)
            if isinstance(value, (str, int, float, bool)) or value is None:
                metrics[key] = value
        return metrics

    def extract_status_percent(self, status_data: Dict[str, Any]) -> Optional[float]:
        for key in ("percent", "percentage", "progress"):
            value = status_data.get(key)
            if isinstance(value, (int, float)):
                return max(0.0, min(100.0, float(value)))
            if isinstance(value, str):
                try:
                    return max(0.0, min(100.0, float(value.strip().rstrip("%"))))
                except ValueError:
                    pass

        metrics = self.safe_status_metrics(status_data)
        processed = None
        total = None
        for key in ("processed", "processedCount", "processedItems", "processedRows", "success", "successCount"):
            value = metrics.get(key)
            if isinstance(value, (int, float)):
                processed = float(value)
                break
        for key in ("total", "totalCount", "totalItems", "totalRows"):
            value = metrics.get(key)
            if isinstance(value, (int, float)) and value > 0:
                total = float(value)
                break
        if processed is not None and total:
            return max(0.0, min(100.0, processed / total * 100.0))
        return None

    def emit_part_progress(
        self,
        phase: str,
        part: Dict[str, Any],
        batch_id: str = "",
        status_name_value: str = "",
        http_status: Any = "",
        elapsed_seconds: Optional[float] = None,
        message: str = "",
        extra: Optional[Dict[str, Any]] = None,
        percent: Optional[float] = None,
    ) -> None:
        self.current_phase = phase
        self.current_batch_id = batch_id or self.current_batch_id
        payload: Dict[str, Any] = {
            "phase": phase,
            "phase_label": self.PHASE_LABELS.get(phase, phase),
            "part_number": int(part.get("part_number", 0)),
            "batch_id": batch_id or "",
            "rows": int(part.get("rows", 0)),
            "bytes": int(part.get("bytes", 0)),
            "status_name": status_name_value,
            "http_status": http_status,
            "completed_parts": self.parts_completed,
            "failed_parts": self.parts_failed,
        }
        if percent is not None:
            payload["percent"] = round(max(0.0, min(100.0, percent)), 2)
        elif phase != "status_polling" or is_success_status(status_name_value) or is_failure_status(status_name_value):
            payload["percent"] = self.progress_percent(phase, status_name_value)
        if elapsed_seconds is not None:
            payload["elapsed_seconds"] = round(elapsed_seconds, 2)
        if message:
            payload["message"] = message
        if extra:
            payload.update(extra)
        emit_progress(payload)

    def append_state(self, row: Dict[str, Any]) -> None:
        append_state_row(self.state_file, row, self.args.encoding)

    def skip_row(self, source_row: int, row: Dict[str, Any], reason: str, detail: str = "") -> None:
        self.reason_counts[reason] += 1
        self.skipped.append({
            "SourceRow": source_row,
            "CodigoSku": row.get("CODIGO SKU", ""),
            "CodigoSucursal": row.get("CODIGO SUCURSAL", ""),
            "Existencia": row.get("EXISTENCIA", ""),
            "Reason": reason,
            "Detail": detail,
        })

    def transform_row(self, source_row: int, row: Dict[str, Any], sku_map: Dict[str, str]) -> Optional[Dict[str, Any]]:
        sku_ref = normalize_sku(row.get("CODIGO SKU", ""))
        if not sku_ref or sku_ref not in sku_map:
            self.skip_row(source_row, row, "MISSING_SKU_ID_MAPPING", sku_ref)
            return None

        item_id = parse_sku_id(sku_map[sku_ref])
        if item_id is None:
            self.skip_row(source_row, row, "INVALID_SKU_ID", str(sku_map.get(sku_ref, "")))
            return None

        container_id = normalize_warehouse(row.get("CODIGO SUCURSAL", ""), self.args.warehouse_mode)
        if not container_id:
            self.skip_row(source_row, row, "EMPTY_WAREHOUSE")
            return None

        quantity = parse_non_negative_int(row.get("EXISTENCIA", ""))
        if quantity is None:
            self.skip_row(source_row, row, "INVALID_QUANTITY", str(row.get("EXISTENCIA", "")))
            return None

        return {
            "item_id": item_id,
            "account_name": self.account,
            "container_id": container_id,
            "quantity": quantity,
            "unlimited": bool_text(self.args.unlimited),
            "lead_time": self.args.lead_time,
            "supply_date": "",
            "seller_id": "",
        }

    def ensure_upload_window(self, expires_at: Optional[datetime]) -> bool:
        if expires_at is None:
            return True
        remaining = (expires_at - datetime.now(timezone.utc)).total_seconds()
        return remaining >= self.args.min_upload_window_seconds

    def create_batch_with_window(self) -> Tuple[int, str, str, str, Dict[str, Any]]:
        if self.client is None:
            raise BatchInventoryError("Cliente VTEX no inicializado")

        last_data: Dict[str, Any] = {}
        last_status = 0
        for _ in range(3):
            last_status, data = self.client.create_batch()
            last_data = data
            batch_id, _method, upload_url, content_type, expires_at = extract_batch_details(data)
            if self.ensure_upload_window(expires_at):
                return last_status, batch_id, upload_url, content_type, data
            print(f"Batch {batch_id} descartado: URL de upload expira demasiado pronto")
        raise BatchInventoryError(
            "No se pudo obtener URL prefirmada con ventana suficiente. "
            f"Ultima respuesta: {format_json(last_data)[:1000]} (HTTP {last_status})"
        )

    def process_part(self, part: Dict[str, Any]) -> None:
        part_number = int(part["part_number"])
        original_path = str(part["path"])
        part_start = time.monotonic()
        batch_id = ""
        final_path = original_path
        self.parts_started += 1

        if self.args.resume and part_number in self.resume_done_parts:
            try:
                os.remove(original_path)
            except OSError:
                pass
            self.parts.append({**part, "status": "RESUME_SKIPPED", "batch_id": "", "path": ""})
            print(f"Parte {part_number:04d}: omitida por --resume (DONE en state file)")
            return

        if self.args.dry_run:
            self.emit_part_progress("dry_run", part)
            batch_id = f"DRYRUN_part{part_number:04d}"
            final_path = os.path.join(self.parts_dir, f"{batch_id}.csv")
            os.replace(original_path, final_path)
            duration = time.monotonic() - part_start
            record = {
                "PartNumber": part_number,
                "BatchId": batch_id,
                "PartFile": final_path,
                "Rows": part["rows"],
                "Bytes": os.path.getsize(final_path),
                "CreateStatus": "DRY_RUN",
                "UploadStatus": "DRY_RUN",
                "CommitStatus": "DRY_RUN",
                "FinalStatus": "DRY_RUN",
                "DurationSeconds": f"{duration:.2f}",
            }
            self.successful.append(record)
            self.parts.append({**part, "status": "DRY_RUN", "batch_id": batch_id, "path": final_path})
            self.append_state({**record, "Status": "DRY_RUN_DONE", "Phase": "dry_run"})
            self.parts_completed += 1
            self.emit_part_progress(
                "done",
                {**part, "bytes": os.path.getsize(final_path)},
                batch_id=batch_id,
                status_name_value="DRY_RUN",
                elapsed_seconds=duration,
                message="Dry-run generado",
            )
            print(f"Parte {part_number:04d}: dry-run generado {final_path}")
            return

        create_status = ""
        upload_status: Any = ""
        commit_status: Any = ""
        final_status = ""
        phase = "create"
        error_file = ""

        try:
            create_started = time.monotonic()
            self.emit_part_progress("create", part)
            create_code, batch_id, upload_url, content_type, create_data = self.create_batch_with_window()
            self.phase_durations["create"] += time.monotonic() - create_started
            create_status = str(create_code)
            final_path = os.path.join(self.parts_dir, f"{batch_id}_part{part_number:04d}.csv")
            os.replace(original_path, final_path)
            self.append_state({
                "PartNumber": part_number,
                "BatchId": batch_id,
                "PartFile": final_path,
                "Rows": part["rows"],
                "Bytes": os.path.getsize(final_path),
                "Status": "CREATED",
                "Phase": "create",
                "StatusCode": create_code,
            })

            phase = "upload"
            part_with_file = {**part, "bytes": os.path.getsize(final_path)}
            self.emit_part_progress("upload", part_with_file, batch_id=batch_id, http_status=create_code)
            upload_started = time.monotonic()
            upload_status, upload_text = upload_file_to_presigned_url(
                final_path,
                upload_url,
                content_type,
                timeout=self.args.timeout,
            )
            self.phase_durations["upload"] += time.monotonic() - upload_started
            if upload_status not in (200, 204):
                raise BatchInventoryError(f"upload fallo HTTP {upload_status}: {upload_text[:1000]}")
            self.append_state({
                "PartNumber": part_number,
                "BatchId": batch_id,
                "PartFile": final_path,
                "Rows": part["rows"],
                "Bytes": os.path.getsize(final_path),
                "Status": "UPLOADED",
                "Phase": "upload",
                "StatusCode": upload_status,
            })

            phase = "commit"
            self.emit_part_progress("commit", part_with_file, batch_id=batch_id, http_status=upload_status)
            commit_started = time.monotonic()
            commit_status, commit_data, commit_text = self.client.commit_batch(batch_id) if self.client else (0, {}, "")
            self.phase_durations["commit"] += time.monotonic() - commit_started
            if commit_status not in (200, 201, 202, 204):
                raise BatchInventoryError(f"commit fallo HTTP {commit_status}: {commit_text[:1000]}")
            self.append_state({
                "PartNumber": part_number,
                "BatchId": batch_id,
                "PartFile": final_path,
                "Rows": part["rows"],
                "Bytes": os.path.getsize(final_path),
                "Status": "COMMITTED",
                "Phase": "commit",
                "StatusCode": commit_status,
            })

            phase = "status"
            self.emit_part_progress("status_polling", part_with_file, batch_id=batch_id, http_status=commit_status)

            def on_status(
                status_code: int,
                status_data: Dict[str, Any],
                name: str,
                elapsed_seconds: float,
                attempt: int,
            ) -> None:
                self.emit_part_progress(
                    "status_polling",
                    part_with_file,
                    batch_id=batch_id,
                    status_name_value=name,
                    http_status=status_code,
                    elapsed_seconds=elapsed_seconds,
                    extra={
                        "attempt": attempt,
                        "status_metrics": self.safe_status_metrics(status_data),
                    },
                    percent=self.extract_status_percent(status_data),
                )

            status_code, status_data, final_status, status_duration = poll_status(
                self.client,
                batch_id,
                poll_interval=self.args.poll_interval,
                max_wait_minutes=self.args.max_status_wait_minutes,
                on_status=on_status,
            )
            self.phase_durations["status"] += status_duration
            is_success = is_success_status(final_status)

            if not is_success:
                error_url = find_error_url(status_data)
                if error_url:
                    error_file = os.path.join(self.errors_dir, f"{batch_id}_errors.csv")
                    try:
                        download_error_file(error_url, error_file, self.args.timeout)
                    except Exception as exc:
                        error_file = os.path.join(self.errors_dir, f"{batch_id}_status_error.json")
                        with open(error_file, "w", encoding="utf-8") as f:
                            json.dump({"download_error": str(exc), "status": status_data}, f, ensure_ascii=False, indent=2)
                else:
                    error_file = os.path.join(self.errors_dir, f"{batch_id}_status_error.json")
                    with open(error_file, "w", encoding="utf-8") as f:
                        json.dump(status_data, f, ensure_ascii=False, indent=2)

                raise BatchInventoryError(f"status terminal no exitoso {final_status}: {format_json(status_data)[:1000]}")

            duration = time.monotonic() - part_start
            record = {
                "PartNumber": part_number,
                "BatchId": batch_id,
                "PartFile": final_path,
                "Rows": part["rows"],
                "Bytes": os.path.getsize(final_path),
                "CreateStatus": create_status,
                "UploadStatus": upload_status,
                "CommitStatus": commit_status,
                "FinalStatus": final_status,
                "DurationSeconds": f"{duration:.2f}",
            }
            self.successful.append(record)
            self.parts.append({**part, "status": "DONE", "batch_id": batch_id, "path": final_path})
            self.append_state({
                **record,
                "Status": "DONE",
                "Phase": "status",
                "StatusCode": status_code,
            })
            self.parts_completed += 1
            self.emit_part_progress(
                "done",
                part_with_file,
                batch_id=batch_id,
                status_name_value=final_status,
                http_status=status_code,
                elapsed_seconds=duration,
                message="Parte completada",
            )
            print(f"Parte {part_number:04d}: DONE batch={batch_id} rows={part['rows']}")

        except Exception as exc:
            duration = time.monotonic() - part_start
            if not error_file and final_status and not is_success_status(final_status):
                error_file = os.path.join(self.errors_dir, f"{batch_id or 'unknown'}_status_error.json")
            record = {
                "PartNumber": part_number,
                "BatchId": batch_id,
                "PartFile": final_path,
                "Rows": part["rows"],
                "Bytes": os.path.getsize(final_path) if os.path.isfile(final_path) else part.get("bytes", 0),
                "Phase": phase,
                "StatusCode": upload_status or commit_status or create_status,
                "FinalStatus": final_status,
                "Error": str(exc),
                "ErrorFile": error_file,
                "DurationSeconds": f"{duration:.2f}",
            }
            self.failed.append(record)
            self.parts.append({**part, "status": "FAILED", "batch_id": batch_id, "path": final_path})
            self.append_state({
                **record,
                "Status": "FAILED",
            })
            self.parts_failed += 1
            self.emit_part_progress(
                "failed",
                part,
                batch_id=batch_id,
                status_name_value=final_status,
                http_status=upload_status or commit_status or create_status,
                elapsed_seconds=duration,
                message=str(exc),
            )
            print(f"Parte {part_number:04d}: FAILED fase={phase} batch={batch_id or 'N/A'} error={exc}")

    def process_input(self, sku_map: Dict[str, str]) -> None:
        max_bytes = int(self.args.max_part_mb * 1024 * 1024)
        writer = PartWriter(self.parts_dir, max_bytes, "utf-8")

        with open(self.args.input_csv, encoding=self.args.encoding, errors="replace", newline="") as f:
            reader = csv.DictReader(f)
            fieldnames = reader.fieldnames or []
            validate_required_columns(fieldnames, ERP_REQUIRED_COLUMNS, self.args.input_csv)

            for source_row, row in enumerate(reader, start=2):
                self.rows_read += 1
                transformed = self.transform_row(source_row, row, sku_map)
                if transformed is None:
                    continue

                result = writer.write_row(transformed)
                if result and result.get("row_too_large"):
                    self.skip_row(
                        source_row,
                        row,
                        "ROW_TOO_LARGE",
                        f"{result.get('row_bytes')} bytes supera max-part-mb",
                    )
                    continue
                self.rows_mapped += 1

                if result:
                    self.process_part(result)

        final_part = writer.finish()
        if final_part:
            self.process_part(final_part)

    def write_outputs(self) -> Dict[str, str]:
        success_path = os.path.join(self.output_dir, f"{self.timestamp}_batch_inventory_successful.csv")
        failed_path = os.path.join(self.output_dir, f"{self.timestamp}_batch_inventory_failed.csv")
        skipped_path = os.path.join(self.output_dir, f"{self.timestamp}_batch_inventory_skipped.csv")
        report_path = os.path.join(self.output_dir, f"{self.timestamp}_batch_inventory_REPORT.md")

        write_csv(success_path, SUCCESS_FIELDNAMES, self.successful, self.args.encoding)
        write_csv(failed_path, FAILED_FIELDNAMES, self.failed, self.args.encoding)
        write_csv(skipped_path, SKIPPED_FIELDNAMES, self.skipped, self.args.encoding)
        self.write_report(report_path, success_path, failed_path, skipped_path)
        return {
            "success": success_path,
            "failed": failed_path,
            "skipped": skipped_path,
            "report": report_path,
        }

    def write_report(self, path: str, success_path: str, failed_path: str, skipped_path: str) -> None:
        ended_at = datetime.now()
        duration = ended_at - self.started_at
        total_bytes = sum(int(part.get("bytes", 0)) for part in self.parts)

        with open(path, "w", encoding="utf-8") as f:
            f.write("# Reporte batch inventory VTEX\n\n")
            f.write(f"**Inicio:** {self.started_at.isoformat(timespec='seconds')}\n\n")
            f.write(f"**Fin:** {ended_at.isoformat(timespec='seconds')}\n\n")
            f.write(f"**Duracion:** {duration}\n\n")
            f.write(f"**Input ERP:** {self.args.input_csv}\n\n")
            f.write(f"**SKU map:** {self.args.sku_map}\n\n")
            f.write(f"**Cuenta/ambiente:** {self.account}/{self.environment}\n\n")
            f.write(f"**Dry-run:** {self.args.dry_run}\n\n")
            f.write("## Resumen\n\n")
            f.write(f"- Filas leidas: {self.rows_read:,}\n")
            f.write(f"- Filas mapeadas: {self.rows_mapped:,}\n")
            f.write(f"- Filas omitidas: {len(self.skipped):,}\n")
            f.write(f"- Partes procesadas: {len(self.parts):,}\n")
            f.write(f"- Bytes en partes: {total_bytes:,}\n")
            f.write(f"- Partes exitosas: {len(self.successful):,}\n")
            f.write(f"- Partes fallidas: {len(self.failed):,}\n")
            f.write(f"- State file: {self.state_file}\n")
            f.write("\n## Omitidos por razon\n\n")
            if self.reason_counts:
                for reason, count in sorted(self.reason_counts.items()):
                    f.write(f"- {reason}: {count:,}\n")
            else:
                f.write("- Ninguno\n")

            f.write("\n## Partes\n\n")
            if self.parts:
                f.write("| Parte | BatchId | Estado | Filas | Bytes | Archivo |\n")
                f.write("| --- | --- | --- | ---: | ---: | --- |\n")
                for part in self.parts:
                    f.write(
                        f"| {int(part.get('part_number', 0)):04d} "
                        f"| {part.get('batch_id', '')} "
                        f"| {part.get('status', '')} "
                        f"| {part.get('rows', 0)} "
                        f"| {part.get('bytes', 0)} "
                        f"| {part.get('path', '')} |\n"
                    )
            else:
                f.write("No se generaron partes.\n")

            f.write("\n## Duracion por fase\n\n")
            if self.phase_durations:
                for phase, seconds in sorted(self.phase_durations.items()):
                    f.write(f"- {phase}: {seconds:.2f}s\n")
            else:
                f.write("- Sin fases API ejecutadas\n")

            f.write("\n## Archivos exportados\n\n")
            f.write(f"- Exitosos: {success_path}\n")
            f.write(f"- Fallidos: {failed_path}\n")
            f.write(f"- Omitidos: {skipped_path}\n")
            if self.failed:
                f.write("\n## Errores descargados\n\n")
                for row in self.failed:
                    if row.get("ErrorFile"):
                        f.write(f"- Batch {row.get('BatchId')}: {row.get('ErrorFile')}\n")

    def run(self) -> Dict[str, str]:
        transform_started = time.monotonic()
        sku_map = load_sku_map(
            self.args.sku_map,
            self.args.sku_ref_column,
            self.args.sku_id_column,
            self.args.encoding,
            self.args.sku_map_header_row,
            self.output_dir,
            self.timestamp,
        )
        print(f"SKU map cargado: {len(sku_map):,} referencias")
        print(f"Output dir: {self.output_dir}")
        print(f"Max parte: {self.args.max_part_mb} MB")
        if self.args.dry_run:
            print("DRY-RUN: no se crearan batches ni se subiran archivos")
        self.process_input(sku_map)
        self.phase_durations["transform"] += time.monotonic() - transform_started
        outputs = self.write_outputs()
        print("Proceso finalizado")
        print(f"Filas leidas={self.rows_read:,} mapeadas={self.rows_mapped:,} omitidas={len(self.skipped):,}")
        print(f"Partes exitosas={len(self.successful):,} fallidas={len(self.failed):,}")
        print(f"Reporte: {outputs['report']}")
        return outputs


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Genera CSV VTEX batch desde ERP Homesentry y lo sube por Logistics batch. "
            "--sku-map es obligatorio porque CODIGO SKU es referencia y VTEX espera item_id entero."
        )
    )
    parser.add_argument("input_csv", help="CSV ERP con columnas CODIGO SKU, CODIGO SUCURSAL, EXISTENCIA")
    parser.add_argument(
        "--sku-map",
        required=True,
        help="Archivo .csv/.xlsx obligatorio con el mapa referencia SKU -> SkuId",
    )
    parser.add_argument("--sku-ref-column", default=DEFAULT_SKU_REF_COLUMN, help="Columna referencia en --sku-map")
    parser.add_argument("--sku-id-column", default=DEFAULT_SKU_ID_COLUMN, help="Columna SkuId entero en --sku-map")
    parser.add_argument(
        "--sku-map-header-row",
        type=int,
        default=2,
        help="Fila de encabezados para --sku-map .xlsx; por defecto 2",
    )
    parser.add_argument(
        "--warehouse-mode",
        choices=["zfill3", "raw"],
        default="zfill3",
        help="Normalizacion de CODIGO SUCURSAL; zfill3 convierte 1->001 y 95->095",
    )
    parser.add_argument("--max-part-mb", type=float, default=450.0, help="Tamano maximo por parte CSV en MB")
    parser.add_argument("--lead-time", default="1.00:00:00", help="Valor lead_time para VTEX batch")
    parser.add_argument("--unlimited", action="store_true", help="Enviar unlimited=true; por defecto false")
    parser.add_argument("--dry-run", action="store_true", help="Genera partes y reportes sin requests a VTEX")
    parser.add_argument(
        "--output-dir",
        default=os.path.join(SCRIPT_DIR, "output"),
        help="Directorio de salida; por defecto output/ junto a este script",
    )
    parser.add_argument("--encoding", default="utf-8-sig", help="Encoding para CSV de entrada/salida")
    parser.add_argument("--poll-interval", type=float, default=10.0, help="Segundos entre consultas de status")
    parser.add_argument(
        "--max-status-wait-minutes",
        type=float,
        default=60.0,
        help="Minutos maximos de espera por status terminal por parte",
    )
    parser.add_argument("--timeout", type=int, default=60, help="Timeout HTTP por request en segundos")
    parser.add_argument(
        "--min-upload-window-seconds",
        type=int,
        default=300,
        help="Ventana minima restante de URL prefirmada antes de subir",
    )
    parser.add_argument("--resume", dest="resume", action="store_true", default=True, help="Saltar partes DONE del state file")
    parser.add_argument("--no-resume", dest="resume", action="store_false", help="Reprocesar aunque exista state file")
    return parser


def main() -> int:
    load_project_env(os.path.join(PROJECT_ROOT, ".env"))
    parser = build_parser()
    args = parser.parse_args()

    if not os.path.isfile(args.input_csv):
        print(f"ERROR: No se encontro input_csv: {args.input_csv}", file=sys.stderr)
        return 1
    if args.max_part_mb <= 0:
        print("ERROR: --max-part-mb debe ser mayor a 0", file=sys.stderr)
        return 1
    if args.poll_interval <= 0:
        print("ERROR: --poll-interval debe ser mayor a 0", file=sys.stderr)
        return 1
    if args.max_status_wait_minutes <= 0:
        print("ERROR: --max-status-wait-minutes debe ser mayor a 0", file=sys.stderr)
        return 1
    if args.sku_map_header_row < 1:
        print("ERROR: --sku-map-header-row debe ser mayor o igual a 1", file=sys.stderr)
        return 1

    try:
        uploader = BatchInventoryUploader(args)
        uploader.run()
    except KeyboardInterrupt:
        print("\nWARN: Proceso interrumpido por el usuario", file=sys.stderr)
        return 130
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
