#!/usr/bin/env python3
"""
delete_sku_files_by_refid.py

Elimina archivos de SKUs en VTEX a partir de codigos de referencia.

Flujo:
- Lee un archivo maestro con referenceCode -> skuId.
- Lee un segundo archivo con referenceCodes objetivo.
- Cruza referencias para obtener los skuId.
- Consulta archivos del SKU con GET /api/catalog/pvt/stockkeepingunit/{skuId}/file.
- Elimina cada archivo con DELETE /api/catalog/pvt/stockkeepingunit/{skuId}/file/{skuFileId}.

Uso:
    python3 delete_sku_files_by_refid.py skus.csv referencias.csv --dry-run
    python3 delete_sku_files_by_refid.py skus.xlsx referencias.csv --dry-run
    python3 delete_sku_files_by_refid.py skus.json referencias.json --delay 1.0
    python3 delete_sku_files_by_refid.py skus.csv refs.csv --mapping-ref-column RefId --mapping-sku-column Id

Requisitos:
- requests
- python-dotenv
- openpyxl para leer archivos .xlsx
- .env en la raiz del proyecto con credenciales VTEX
"""

import argparse
import csv
import json
import os
import sys
import time
from datetime import datetime

import requests
from dotenv import load_dotenv


PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ENV_PATH = os.path.join(PROJECT_ROOT, ".env")
load_dotenv(dotenv_path=ENV_PATH)

RETRY_STATUS_CODES = (429, 500, 502, 503, 504)
DEFAULT_OUTPUT_PREFIX = "sku_file_deletion"
MAPPING_REF_ALIASES = [
    "RefId",
    "refId",
    "referenceCode",
    "ReferenceCode",
    "CODIGO SKU",
    "Código Referencia",
    "Codigo Referencia",
    "CodigoReferencia",
    "_SKUReferenceCode",
    "SKU reference code",
    "SKU Reference Code",
    "SKU",
]
MAPPING_SKU_ALIASES = ["SkuId", "skuId", "Id", "id", "_SkuId", "SKU ID", "Sku ID", "sku id"]
REFERENCE_ALIASES = MAPPING_REF_ALIASES
FILE_ID_ALIASES = ["Id", "id", "SkuFileId", "skuFileId", "FileId", "fileId"]
FILE_NAME_ALIASES = ["Name", "name", "FileName", "fileName", "Label", "label"]
FILE_URL_ALIASES = ["Url", "url", "FileUrl", "fileUrl", "ImageUrl", "imageUrl"]


class FatalInputError(Exception):
    """Error de entrada o configuracion que impide continuar."""


class VTEXRequestError(Exception):
    """Error HTTP o de red contra VTEX."""

    def __init__(self, stage, status_code, message):
        super(VTEXRequestError, self).__init__(message)
        self.stage = stage
        self.status_code = status_code
        self.message = message


def parse_args():
    parser = argparse.ArgumentParser(
        description="Eliminar archivos de SKUs VTEX cruzando referenceCode -> skuId",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos:
  python3 delete_sku_files_by_refid.py skus.xlsx referencias.csv --dry-run --delay 1.0
  python3 delete_sku_files_by_refid.py skus.csv referencias.csv --delay 1.0 --output-prefix borrado_imagenes
  python3 delete_sku_files_by_refid.py skus.json refs.json --mapping-ref-column RefId --mapping-sku-column Id
        """,
    )
    parser.add_argument("mapping_file", help="CSV/XLSX/JSON maestro con referenceCode y skuId")
    parser.add_argument("references_file", help="CSV/JSON con referenceCodes objetivo")
    parser.add_argument("--dry-run", action="store_true",
                        help="Consulta archivos reales, pero no ejecuta DELETE")
    parser.add_argument("--delay", type=float, default=1.0,
                        help="Segundos entre requests GET/DELETE (default: 1.0)")
    parser.add_argument("--timeout", type=int, default=30,
                        help="Timeout de requests en segundos (default: 30)")
    parser.add_argument("--limit", type=int, default=None,
                        help="Maximo de referencias objetivo a procesar tras deduplicar")
    parser.add_argument("--mapping-ref-column", default=None,
                        help="Columna/campo del maestro con el referenceCode")
    parser.add_argument("--mapping-sku-column", default=None,
                        help="Columna/campo del maestro con el skuId VTEX")
    parser.add_argument("--references-column", default=None,
                        help="Columna/campo del archivo objetivo con referenceCode")
    parser.add_argument("--account", default=None,
                        help="Cuenta VTEX (sobrescribe VTEX_ACCOUNT_NAME)")
    parser.add_argument("--environment", default=None,
                        help="Environment VTEX (sobrescribe VTEX_ENVIRONMENT)")
    parser.add_argument("--output-prefix", default=DEFAULT_OUTPUT_PREFIX,
                        help="Prefijo de archivos de salida (default: sku_file_deletion)")
    parser.add_argument("--retries", type=int, default=3,
                        help="Reintentos para HTTP 429/5xx (default: 3)")
    return parser.parse_args()


def load_credentials(account_override=None, environment_override=None):
    app_key = os.getenv("X-VTEX-API-AppKey")
    app_token = os.getenv("X-VTEX-API-AppToken")
    account = account_override or os.getenv("VTEX_ACCOUNT_NAME")
    environment = environment_override or os.getenv("VTEX_ENVIRONMENT", "vtexcommercestable")

    missing = []
    if not app_key:
        missing.append("X-VTEX-API-AppKey")
    if not app_token:
        missing.append("X-VTEX-API-AppToken")
    if not account:
        missing.append("VTEX_ACCOUNT_NAME")
    if not environment:
        missing.append("VTEX_ENVIRONMENT")
    if missing:
        raise FatalInputError("Credenciales VTEX faltantes: {}".format(", ".join(missing)))

    return app_key, app_token, account, environment


def build_base_url(account, environment):
    env = str(environment).strip()
    if env.startswith("http://") or env.startswith("https://"):
        return env.rstrip("/")
    if env.endswith(".com.br"):
        return "https://{}.{}".format(account, env)
    return "https://{}.{}.com.br".format(account, env)


def normalize_code(value):
    if value is None:
        return ""
    code = str(value).strip()
    return code


def load_tabular_or_json(path):
    if not os.path.exists(path):
        raise FatalInputError("No se encontro el archivo '{}'".format(path))

    extension = os.path.splitext(path)[1].lower()
    if extension == ".csv":
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            return rows, reader.fieldnames or []

    if extension == ".json":
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data, infer_fieldnames(data)

    if extension == ".xlsx":
        return load_xlsx(path, header_row=2)

    raise FatalInputError("Formato no soportado para '{}'. Use .csv, .xlsx o .json".format(path))


def load_xlsx(path, header_row=2):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise FatalInputError("Para leer .xlsx instale openpyxl: pip install openpyxl")

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active

    header_values = next(
        sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True),
        None,
    )
    if not header_values:
        raise FatalInputError("El archivo .xlsx no tiene encabezados en la fila {}".format(header_row))

    fieldnames = []
    for index, value in enumerate(header_values, start=1):
        header = normalize_code(value)
        fieldnames.append(header if header else "column_{}".format(index))

    rows = []
    for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        row = {}
        has_value = False
        for field, value in zip(fieldnames, values):
            cell_value = normalize_code(value)
            row[field] = cell_value
            if cell_value:
                has_value = True
        if has_value:
            rows.append(row)

    workbook.close()
    return rows, fieldnames


def infer_fieldnames(data):
    if isinstance(data, list):
        for item in data:
            if isinstance(item, dict):
                return list(item.keys())
        return []
    if isinstance(data, dict):
        dict_values = [value for value in data.values() if isinstance(value, dict)]
        if dict_values:
            fieldnames = set()
            for item in dict_values:
                fieldnames.update(item.keys())
            return list(fieldnames)
        return list(data.keys())
    return []


def detect_column(fieldnames, explicit_name, aliases, label):
    available = fieldnames or []
    if explicit_name:
        if explicit_name in available:
            return explicit_name
        raise FatalInputError(
            "Columna '{}' no encontrada para {}. Columnas disponibles: {}".format(
                explicit_name, label, ", ".join(available) or "(sin columnas)"
            )
        )

    for alias in aliases:
        if alias in available:
            return alias

    lower_map = {str(name).strip().lower(): name for name in available}
    for alias in aliases:
        found = lower_map.get(alias.lower())
        if found:
            return found

    raise FatalInputError(
        "No se pudo detectar columna para {}. Columnas disponibles: {}".format(
            label, ", ".join(available) or "(sin columnas)"
        )
    )


def rows_from_mapping_data(data):
    if isinstance(data, list):
        return [row for row in data if isinstance(row, dict)]
    if isinstance(data, dict):
        rows = []
        for key, value in data.items():
            if isinstance(value, dict):
                row = dict(value)
                row.setdefault("referenceCode", key)
                rows.append(row)
            else:
                rows.append({"referenceCode": key, "skuId": value})
        return rows
    raise FatalInputError("El archivo maestro debe ser CSV, lista JSON o diccionario JSON")


def read_mapping(mapping_file, ref_column=None, sku_column=None):
    data, fieldnames = load_tabular_or_json(mapping_file)
    rows = rows_from_mapping_data(data)
    if not rows:
        raise FatalInputError("El archivo maestro no contiene filas validas")

    fieldnames = fieldnames or infer_fieldnames(rows)
    ref_col = detect_column(fieldnames, ref_column, MAPPING_REF_ALIASES, "referenceCode del maestro")
    sku_col = detect_column(fieldnames, sku_column, MAPPING_SKU_ALIASES, "skuId del maestro")

    mapping = {}
    conflicts = []
    conflict_refs = set()
    invalid_rows = []

    for index, row in enumerate(rows, start=1):
        ref_code = normalize_code(row.get(ref_col))
        sku_id = normalize_code(row.get(sku_col))
        if not ref_code or not sku_id:
            invalid_rows.append({
                "referenceCode": ref_code,
                "skuId": sku_id,
                "skuFileId": "",
                "stage": "mapping",
                "statusCode": "",
                "error": "Fila {} sin referenceCode o skuId".format(index),
                "timestamp": datetime.now().isoformat(timespec="seconds"),
            })
            continue

        existing = mapping.get(ref_code)
        if ref_code in conflict_refs:
            continue
        if existing and existing != sku_id:
            conflicts.append({
                "referenceCode": ref_code,
                "skuId": "{} | {}".format(existing, sku_id),
                "skuFileId": "",
                "stage": "mapping",
                "statusCode": "",
                "error": "Conflicto: la misma referencia apunta a dos skuId distintos",
                "timestamp": datetime.now().isoformat(timespec="seconds"),
            })
            conflict_refs.add(ref_code)
            mapping.pop(ref_code, None)
            continue

        if not existing:
            mapping[ref_code] = sku_id

    return mapping, conflicts, invalid_rows, ref_col, sku_col


def read_reference_targets(references_file, references_column=None):
    data, fieldnames = load_tabular_or_json(references_file)
    references = []

    if isinstance(data, list):
        if all(not isinstance(item, dict) for item in data):
            references = [normalize_code(item) for item in data]
        else:
            rows = [row for row in data if isinstance(row, dict)]
            if not rows:
                raise FatalInputError("El archivo objetivo no contiene objetos validos")
            fieldnames = fieldnames or infer_fieldnames(rows)
            ref_col = detect_column(fieldnames, references_column, REFERENCE_ALIASES, "referenceCode objetivo")
            references = [normalize_code(row.get(ref_col)) for row in rows]
    elif isinstance(data, dict):
        if references_column:
            rows = rows_from_mapping_data(data)
            fieldnames = fieldnames or infer_fieldnames(rows)
            ref_col = detect_column(fieldnames, references_column, REFERENCE_ALIASES, "referenceCode objetivo")
            references = [normalize_code(row.get(ref_col)) for row in rows]
        else:
            references = [normalize_code(key) for key in data.keys()]
    else:
        raise FatalInputError("El archivo objetivo debe ser CSV, lista JSON o diccionario JSON")

    seen = set()
    unique = []
    total_non_empty = 0
    for ref_code in references:
        if not ref_code:
            continue
        total_non_empty += 1
        if ref_code not in seen:
            seen.add(ref_code)
            unique.append(ref_code)

    return unique, total_non_empty


def short_body(response):
    text = response.text or ""
    return text[:800].replace("\n", " ").replace("\r", " ")


def value_from_aliases(record, aliases):
    if not isinstance(record, dict):
        return ""
    for alias in aliases:
        value = normalize_code(record.get(alias))
        if value:
            return value
    lower_map = {str(key).lower(): key for key in record.keys()}
    for alias in aliases:
        key = lower_map.get(alias.lower())
        if key is not None:
            value = normalize_code(record.get(key))
            if value:
                return value
    return ""


def summarize_record(record):
    try:
        return json.dumps(record, ensure_ascii=False)[:500]
    except Exception:
        return str(record)[:500]


class VTEXClient:
    def __init__(self, base_url, app_key, app_token, timeout=30, retries=3):
        self.base_url = base_url.rstrip("/")
        self.timeout = timeout
        self.retries = max(0, retries)
        self.session = requests.Session()
        self.session.headers.update({
            "Accept": "application/json",
            "Content-Type": "application/json",
            "X-VTEX-API-AppKey": app_key,
            "X-VTEX-API-AppToken": app_token,
        })

    def _request(self, method, path, stage, **kwargs):
        url = "{}{}".format(self.base_url, path)
        attempts = self.retries + 1
        wait = 1.0

        for attempt in range(1, attempts + 1):
            try:
                response = self.session.request(method, url, timeout=self.timeout, **kwargs)
            except requests.exceptions.Timeout:
                if attempt < attempts:
                    time.sleep(wait)
                    wait *= 2
                    continue
                raise VTEXRequestError(stage, "", "Timeout despues de {} intentos".format(attempts))
            except requests.exceptions.RequestException as exc:
                if attempt < attempts:
                    time.sleep(wait)
                    wait *= 2
                    continue
                raise VTEXRequestError(stage, "", "RequestException: {}".format(exc))

            if response.status_code in RETRY_STATUS_CODES and attempt < attempts:
                time.sleep(wait)
                wait *= 2
                continue

            return response

        raise VTEXRequestError(stage, "", "Request fallido despues de {} intentos".format(attempts))

    def list_sku_files(self, sku_id):
        path = "/api/catalog/pvt/stockkeepingunit/{}/file".format(sku_id)
        response = self._request("GET", path, "list")
        if response.status_code >= 400:
            raise VTEXRequestError("list", response.status_code, "HTTP {}: {}".format(response.status_code, short_body(response)))
        try:
            data = response.json()
        except ValueError:
            raise VTEXRequestError("list", response.status_code, "Respuesta GET no es JSON valido")

        if isinstance(data, list):
            return data
        if isinstance(data, dict) and isinstance(data.get("data"), list):
            return data.get("data")
        return []

    def delete_sku_file(self, sku_id, sku_file_id):
        path = "/api/catalog/pvt/stockkeepingunit/{}/file/{}".format(sku_id, sku_file_id)
        response = self._request("DELETE", path, "delete")
        if response.status_code in (200, 204):
            return response.status_code, "deleted", ""
        if response.status_code == 404:
            return response.status_code, "not_found", ""
        return response.status_code, "error", "HTTP {}: {}".format(response.status_code, short_body(response))


def extract_file_id(file_record):
    return value_from_aliases(file_record, FILE_ID_ALIASES)


def build_deletion_plan(mapping, references, conflicts):
    conflict_refs = set(item["referenceCode"] for item in conflicts)
    plan = []
    errors = []

    for ref_code in references:
        if ref_code in conflict_refs:
            errors.append({
                "referenceCode": ref_code,
                "skuId": "",
                "skuFileId": "",
                "stage": "mapping",
                "statusCode": "",
                "error": "Referencia con conflicto en archivo maestro; no se procesa",
                "timestamp": datetime.now().isoformat(timespec="seconds"),
            })
            continue

        sku_id = mapping.get(ref_code)
        if not sku_id:
            errors.append({
                "referenceCode": ref_code,
                "skuId": "",
                "skuFileId": "",
                "stage": "mapping",
                "statusCode": "",
                "error": "Referencia no encontrada en archivo maestro",
                "timestamp": datetime.now().isoformat(timespec="seconds"),
            })
            continue

        plan.append({"referenceCode": ref_code, "skuId": sku_id})

    return plan, errors


def result_row(reference_code, sku_id, sku_file_id, file_name, file_url, action, status_code, result, error):
    return {
        "referenceCode": reference_code,
        "skuId": sku_id,
        "skuFileId": sku_file_id,
        "fileName": file_name,
        "fileUrl": file_url,
        "action": action,
        "statusCode": status_code,
        "result": result,
        "error": error,
        "timestamp": datetime.now().isoformat(timespec="seconds"),
    }


def error_row(reference_code, sku_id, sku_file_id, stage, status_code, error):
    return {
        "referenceCode": reference_code,
        "skuId": sku_id,
        "skuFileId": sku_file_id,
        "stage": stage,
        "statusCode": status_code,
        "error": error,
        "timestamp": datetime.now().isoformat(timespec="seconds"),
    }


def sleep_between_requests(delay):
    if delay and delay > 0:
        time.sleep(delay)


def process_deletions(client, deletion_plan, args):
    results = []
    errors = []
    stats = {
        "skus_consulted": 0,
        "skus_no_files": 0,
        "files_found": 0,
        "files_deleted": 0,
        "files_simulated": 0,
        "files_not_found": 0,
        "api_errors": 0,
    }
    total = len(deletion_plan)
    start_time = datetime.now()

    for index, item in enumerate(deletion_plan, start=1):
        ref_code = item["referenceCode"]
        sku_id = item["skuId"]

        try:
            files = client.list_sku_files(sku_id)
            stats["skus_consulted"] += 1
        except VTEXRequestError as exc:
            stats["api_errors"] += 1
            errors.append(error_row(ref_code, sku_id, "", exc.stage, exc.status_code, exc.message))
            results.append(result_row(ref_code, sku_id, "", "", "", "list", exc.status_code, "error", exc.message))
            print("[{}/{}] {} -> error consultando archivos ({})".format(index, total, ref_code, exc.message))
            sleep_between_requests(args.delay)
            continue

        sleep_between_requests(args.delay)

        if not files:
            stats["skus_no_files"] += 1
            results.append(result_row(ref_code, sku_id, "", "", "", "list", 200, "no_files", ""))
            print("[{}/{}] {} -> SKU {} sin archivos".format(index, total, ref_code, sku_id))
        else:
            stats["files_found"] += len(files)
            print("[{}/{}] {} -> SKU {} con {} archivo(s)".format(index, total, ref_code, sku_id, len(files)))

        for file_record in files:
            sku_file_id = extract_file_id(file_record)
            file_name = value_from_aliases(file_record, FILE_NAME_ALIASES)
            file_url = value_from_aliases(file_record, FILE_URL_ALIASES)

            if not sku_file_id:
                message = "No se encontro skuFileId en objeto: {}".format(summarize_record(file_record))
                stats["api_errors"] += 1
                errors.append(error_row(ref_code, sku_id, "", "extract_file_id", "", message))
                results.append(result_row(ref_code, sku_id, "", file_name, file_url, "delete", "", "error", message))
                continue

            if args.dry_run:
                stats["files_simulated"] += 1
                results.append(result_row(ref_code, sku_id, sku_file_id, file_name, file_url, "delete", "", "simulated", ""))
                continue

            try:
                status_code, result, error = client.delete_sku_file(sku_id, sku_file_id)
            except VTEXRequestError as exc:
                stats["api_errors"] += 1
                errors.append(error_row(ref_code, sku_id, sku_file_id, exc.stage, exc.status_code, exc.message))
                results.append(result_row(ref_code, sku_id, sku_file_id, file_name, file_url, "delete", exc.status_code, "error", exc.message))
                sleep_between_requests(args.delay)
                continue

            if result == "deleted":
                stats["files_deleted"] += 1
            elif result == "not_found":
                stats["files_not_found"] += 1
            else:
                stats["api_errors"] += 1
                errors.append(error_row(ref_code, sku_id, sku_file_id, "delete", status_code, error))

            results.append(result_row(ref_code, sku_id, sku_file_id, file_name, file_url, "delete", status_code, result, error))
            sleep_between_requests(args.delay)

        if index % 10 == 0 or index == total:
            elapsed = (datetime.now() - start_time).total_seconds()
            print("Progreso: {}/{} ({:.1f}%) | archivos: {} | errores: {} | {:.0f}s".format(
                index, total, (index / total * 100) if total else 0, stats["files_found"], stats["api_errors"], elapsed
            ))

    return results, errors, stats


def write_csv(path, rows, fieldnames):
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_report(path, args, stats, files, mapping_info):
    mode = "DRY-RUN" if args.dry_run else "REAL"
    lines = [
        "# Reporte de Eliminacion de Archivos SKU VTEX",
        "",
        "**Generado:** {}  ".format(stats["generated_at"]),
        "**Cuenta VTEX:** {}  ".format(stats["account"]),
        "**Environment:** {}  ".format(stats["environment"]),
        "**Modo:** {}".format(mode),
        "",
        "---",
        "",
        "## Resumen",
        "",
        "| Metrica | Cantidad |",
        "|---------|----------|",
        "| Referencias objetivo | {} |".format(stats["target_references"]),
        "| Referencias mapeadas | {} |".format(stats["mapped_references"]),
        "| Referencias sin match/conflicto | {} |".format(stats["unmatched_references"]),
        "| SKUs consultados | {} |".format(stats["skus_consulted"]),
        "| SKUs sin archivos | {} |".format(stats["skus_no_files"]),
        "| Archivos encontrados | {} |".format(stats["files_found"]),
        "| Archivos eliminados | {} |".format(stats["files_deleted"]),
        "| Archivos simulados | {} |".format(stats["files_simulated"]),
        "| Archivos no encontrados (404) | {} |".format(stats["files_not_found"]),
        "| Errores | {} |".format(stats["errors"]),
        "",
        "## Configuracion",
        "",
        "- Mapping: `{}`".format(args.mapping_file),
        "- Referencias: `{}`".format(args.references_file),
        "- Columna referencia maestro: `{}`".format(mapping_info["mapping_ref_column"]),
        "- Columna sku maestro: `{}`".format(mapping_info["mapping_sku_column"]),
        "- Delay: {}s".format(args.delay),
        "- Timeout: {}s".format(args.timeout),
        "- Reintentos: {}".format(args.retries),
        "- Dry-run: {}".format("Si" if args.dry_run else "No"),
        "",
        "## Archivos generados",
        "",
        "- Resultados: `{}`".format(files["results"]),
        "- Reporte: `{}`".format(files["report"]),
    ]

    if files.get("errors"):
        lines.append("- Errores: `{}`".format(files["errors"]))

    lines += [
        "",
        "---",
        "",
        "*Reporte generado automaticamente por `delete_sku_files_by_refid.py`*",
    ]

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def main():
    args = parse_args()

    try:
        app_key, app_token, account, environment = load_credentials(args.account, args.environment)
        base_url = build_base_url(account, environment)
        mapping, conflicts, invalid_mapping_rows, ref_col, sku_col = read_mapping(
            args.mapping_file,
            args.mapping_ref_column,
            args.mapping_sku_column,
        )
        references, total_reference_rows = read_reference_targets(args.references_file, args.references_column)

        if args.limit is not None and args.limit > 0:
            original_count = len(references)
            references = references[:args.limit]
            print("Limite aplicado: {} de {} referencias unicas".format(len(references), original_count))

        if not references:
            raise FatalInputError("No se encontraron referencias objetivo")

        deletion_plan, mapping_errors = build_deletion_plan(mapping, references, conflicts)
        preflight_errors = mapping_errors

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        prefix = args.output_prefix
        results_csv = "{}_results_{}.csv".format(prefix, timestamp)
        errors_csv = "{}_errors_{}.csv".format(prefix, timestamp)
        report_md = "{}_report_{}.md".format(prefix, timestamp)

        print("=" * 70)
        print("VTEX SKU File Deleter by RefId")
        print("=" * 70)
        print("Cuenta: {}".format(account))
        print("Environment: {}".format(environment))
        print("Base URL: {}".format(base_url))
        print("Modo: {}".format("DRY-RUN" if args.dry_run else "REAL"))
        print("Referencias objetivo unicas: {} ({} no vacias antes de deduplicar)".format(len(references), total_reference_rows))
        print("Referencias mapeadas a procesar: {}".format(len(deletion_plan)))
        print("Errores de mapping/preflight para referencias objetivo: {}".format(len(preflight_errors)))
        if conflicts or invalid_mapping_rows:
            print("Avisos ignorados del maestro completo: {} conflictos, {} filas incompletas".format(
                len(conflicts), len(invalid_mapping_rows)
            ))
        print("Delay: {}s | Timeout: {}s | Reintentos: {}".format(args.delay, args.timeout, args.retries))
        print("=" * 70)

        client = VTEXClient(base_url, app_key, app_token, timeout=args.timeout, retries=args.retries)
        results, api_errors, process_stats = process_deletions(client, deletion_plan, args)
        all_errors = preflight_errors + api_errors

        result_fields = [
            "referenceCode",
            "skuId",
            "skuFileId",
            "fileName",
            "fileUrl",
            "action",
            "statusCode",
            "result",
            "error",
            "timestamp",
        ]
        error_fields = ["referenceCode", "skuId", "skuFileId", "stage", "statusCode", "error", "timestamp"]

        write_csv(results_csv, results, result_fields)
        if all_errors:
            write_csv(errors_csv, all_errors, error_fields)

        report_stats = dict(process_stats)
        report_stats.update({
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "account": account,
            "environment": environment,
            "target_references": len(references),
            "mapped_references": len(deletion_plan),
            "unmatched_references": len(preflight_errors),
            "errors": len(all_errors),
        })
        files = {"results": results_csv, "report": report_md}
        if all_errors:
            files["errors"] = errors_csv
        mapping_info = {"mapping_ref_column": ref_col, "mapping_sku_column": sku_col}
        write_report(report_md, args, report_stats, files, mapping_info)

        print("\nProceso completado")
        print("Resultados: {}".format(results_csv))
        if all_errors:
            print("Errores: {}".format(errors_csv))
        print("Reporte: {}".format(report_md))

        return 1 if all_errors else 0

    except FatalInputError as exc:
        print("Error fatal: {}".format(exc))
        return 2


if __name__ == "__main__":
    sys.exit(main())
